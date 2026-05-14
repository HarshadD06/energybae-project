"""
Energybae — Solar Load Calculator
Upload any electricity bill (PDF or image) → AI extracts data → Excel output
Run: python app.py
"""

import anthropic
import base64
import json
import os
import re
from datetime import datetime
from pathlib import Path

import gradio as gr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Anthropic client ──────────────────────────────────────────────────────────
client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))

# ── AI extraction ─────────────────────────────────────────────────────────────

def extract_bill_data(file_path: str) -> dict:
    """Send bill to Claude and extract structured data."""
    path = Path(file_path)
    suffix = path.suffix.lower()

    with open(file_path, "rb") as f:
        raw = f.read()
    b64 = base64.standard_b64encode(raw).decode()

    if suffix == ".pdf":
        content_block = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
        }
    elif suffix in (".jpg", ".jpeg"):
        content_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": b64},
        }
    elif suffix == ".png":
        content_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": b64},
        }
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    prompt = """Extract these fields from this electricity bill.
Return ONLY a valid JSON object — no markdown, no explanation.

{
  "consumer_name": "",
  "consumer_number": "",
  "meter_number": "",
  "discom": "",
  "billing_month": "",
  "units_consumed": 0,
  "sanctioned_load_kw": 0,
  "connected_load_kw": 0,
  "tariff_category": "",
  "supply_voltage": "",
  "total_amount": 0,
  "fixed_charges": 0,
  "energy_charges": 0
}

Rules:
- Numeric fields: numbers only (no units/symbols), null if not found
- units_consumed: total kWh for the billing period
- sanctioned_load_kw: use connected load if sanctioned not available
- billing_month: format as "Mon YYYY" e.g. "Jan 2024"
"""

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1000,
        messages=[
            {
                "role": "user",
                "content": [
                    content_block,
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )

    text = "".join(b.text for b in response.content if hasattr(b, "text"))
    text = re.sub(r"```json|```", "", text).strip()
    return json.loads(text)


# ── Solar calculations ────────────────────────────────────────────────────────

def calculate_solar(units: float, sun_hours: float, efficiency: float,
                    tariff_rate: float, cost_per_kwp: float) -> dict:
    if not units or units <= 0:
        return {}
    eff = efficiency / 100
    daily_units = units / 30
    system_kwp = daily_units / (sun_hours * eff)
    monthly_savings = units * tariff_rate
    annual_savings = monthly_savings * 12
    system_cost = system_kwp * cost_per_kwp
    payback_years = system_cost / annual_savings
    co2_kg_year = units * 12 * 0.82        # India grid: 0.82 kg CO2/kWh
    units_per_year = units * 12
    return {
        "system_kwp": round(system_kwp, 2),
        "daily_units": round(daily_units, 2),
        "monthly_savings": round(monthly_savings),
        "annual_savings": round(annual_savings),
        "system_cost": round(system_cost),
        "payback_years": round(payback_years, 1),
        "co2_kg_year": round(co2_kg_year),
        "units_per_year": round(units_per_year),
    }


# ── Excel generation ──────────────────────────────────────────────────────────

def make_excel(bill: dict, solar_params: dict, calc: dict, out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solar Report"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 28

    # Colours
    green  = "1E6B2E"
    lgreen = "D6EFD8"
    lgray  = "F5F5F5"
    white  = "FFFFFF"

    def hdr(row, text, bg=green, fg=white, sz=11):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color=fg, size=sz)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 20

    def row2(r, label, value, shade=False):
        bg = lgray if shade else white
        la = ws.cell(row=r, column=1, value=label)
        va = ws.cell(row=r, column=2, value=value)
        for c in (la, va):
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="center")
        la.font = Font(size=10)
        va.font = Font(size=10, bold=True)
        ws.row_dimensions[r].height = 18

    thin = Side(style="thin", color="CCCCCC")
    def border_range(r1, c1, r2, c2):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 1
    # Title
    title = ws.cell(row=r, column=1, value="⚡ Energybae — Solar Load Calculator")
    title.font = Font(bold=True, size=14, color=white)
    title.fill = PatternFill("solid", fgColor=green)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 32
    r += 1
    ws.cell(r, 1, f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
    ws.cell(r, 1).font = Font(italic=True, size=9, color="888888")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2

    # Bill data
    hdr(r, "CUSTOMER BILL DATA"); r += 1
    fields = [
        ("Consumer Name",         bill.get("consumer_name")),
        ("Consumer Number",       bill.get("consumer_number")),
        ("Meter Number",          bill.get("meter_number")),
        ("DISCOM / Utility",      bill.get("discom")),
        ("Billing Month",         bill.get("billing_month")),
        ("Tariff Category",       bill.get("tariff_category")),
        ("Supply Voltage",        bill.get("supply_voltage")),
        ("Units Consumed (kWh)",  bill.get("units_consumed")),
        ("Sanctioned Load (kW)",  bill.get("sanctioned_load_kw")),
        ("Total Bill Amount (₹)", bill.get("total_amount")),
        ("Energy Charges (₹)",    bill.get("energy_charges")),
        ("Fixed Charges (₹)",     bill.get("fixed_charges")),
    ]
    for i, (lbl, val) in enumerate(fields):
        row2(r, lbl, val if val not in (None, 0, "") else "—", shade=i % 2 == 0)
        r += 1
    border_range(r - len(fields), 1, r - 1, 2)
    r += 1

    # Solar parameters
    hdr(r, "SOLAR SIZING PARAMETERS"); r += 1
    params = [
        ("Sun Hours / Day",          solar_params.get("sun_hours")),
        ("System Efficiency (%)",    solar_params.get("efficiency")),
        ("Electricity Rate (₹/unit)",solar_params.get("tariff_rate")),
        ("System Cost (₹/kWp)",      solar_params.get("cost_per_kwp")),
    ]
    for i, (lbl, val) in enumerate(params):
        row2(r, lbl, val, shade=i % 2 == 0); r += 1
    border_range(r - len(params), 1, r - 1, 2)
    r += 1

    # Results
    hdr(r, "SOLAR SIZING RESULTS", bg="145A32"); r += 1
    results = [
        ("Recommended System Size (kWp)",  calc.get("system_kwp")),
        ("Daily Generation (kWh/day)",     calc.get("daily_units")),
        ("Annual Units Generated (kWh)",   calc.get("units_per_year")),
        ("Estimated System Cost (₹)",      f"₹ {calc.get('system_cost', 0):,}"),
        ("Monthly Savings (₹)",            f"₹ {calc.get('monthly_savings', 0):,}"),
        ("Annual Savings (₹)",             f"₹ {calc.get('annual_savings', 0):,}"),
        ("Payback Period (years)",         calc.get("payback_years")),
        ("Annual CO₂ Offset (kg)",         f"{calc.get('co2_kg_year', 0):,} kg"),
    ]
    for i, (lbl, val) in enumerate(results):
        row2(r, lbl, val, shade=i % 2 == 0); r += 1
    border_range(r - len(results), 1, r - 1, 2)
    r += 2

    # Notes
    hdr(r, "CALCULATION NOTES", bg="555555"); r += 1
    notes = [
        "System size (kWp) = (monthly units ÷ 30) ÷ (sun hours × efficiency)",
        "CO₂ factor: 0.82 kg per kWh (India CEA grid average)",
        "Savings = monthly units × electricity rate",
        "Payback = system cost ÷ annual savings",
        "Sun hours default: 5.5 hr/day (Maharashtra average)",
    ]
    for note in notes:
        c = ws.cell(row=r, column=1, value=f"  • {note}")
        c.font = Font(size=9, color="555555")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1

    wb.save(out_path)
    return out_path


# ── Gradio UI ─────────────────────────────────────────────────────────────────

def process_bill(file, sun_hours, efficiency, tariff_rate, cost_per_kwp):
    if file is None:
        return ("Please upload a bill file.", "", "", "", "", "", "", "", "", None)

    status = ""
    bill = {}

    # Extract
    try:
        bill = extract_bill_data(file.name)
        status = "✅ Bill data extracted by AI — review below"
    except Exception as e:
        status = f"⚠️ AI extraction failed ({e}). Fill fields manually."

    calc = calculate_solar(
        bill.get("units_consumed") or 0,
        sun_hours, efficiency, tariff_rate, cost_per_kwp
    )

    # Build result summary
    if calc:
        summary = (
            f"🔆 System size: {calc['system_kwp']} kWp\n"
            f"💰 Monthly savings: ₹{calc['monthly_savings']:,}\n"
            f"📅 Payback: {calc['payback_years']} years\n"
            f"🌿 CO₂ offset: {calc['co2_kg_year']:,} kg/year\n"
            f"💵 System cost est.: ₹{calc['system_cost']:,}"
        )
    else:
        summary = "Enter units consumed to calculate solar sizing."

    # Excel
    excel_path = None
    if calc:
        name = (bill.get("consumer_name") or "customer").replace(" ", "_")
        month = (bill.get("billing_month") or "report").replace(" ", "_")
        excel_path = f"/tmp/Energybae_Solar_{name}_{month}.xlsx"
        make_excel(bill, {
            "sun_hours": sun_hours, "efficiency": efficiency,
            "tariff_rate": tariff_rate, "cost_per_kwp": cost_per_kwp
        }, calc, excel_path)

    return (
        status,
        bill.get("consumer_name", ""),
        bill.get("consumer_number", ""),
        bill.get("billing_month", ""),
        bill.get("tariff_category", ""),
        str(bill.get("units_consumed") or ""),
        str(bill.get("sanctioned_load_kw") or ""),
        str(bill.get("total_amount") or ""),
        summary,
        excel_path,
    )


def recalculate(units_str, sun_hours, efficiency, tariff_rate, cost_per_kwp,
                name, consumer_no, month, tariff_cat, load_str, amount_str):
    try:
        units = float(units_str) if units_str else 0
    except ValueError:
        units = 0
    calc = calculate_solar(units, sun_hours, efficiency, tariff_rate, cost_per_kwp)
    if not calc:
        return "Enter units consumed to calculate.", None

    summary = (
        f"🔆 System size: {calc['system_kwp']} kWp\n"
        f"💰 Monthly savings: ₹{calc['monthly_savings']:,}\n"
        f"📅 Payback: {calc['payback_years']} years\n"
        f"🌿 CO₂ offset: {calc['co2_kg_year']:,} kg/year\n"
        f"💵 System cost est.: ₹{calc['system_cost']:,}"
    )

    bill = {
        "consumer_name": name, "consumer_number": consumer_no,
        "billing_month": month, "tariff_category": tariff_cat,
        "units_consumed": units,
        "sanctioned_load_kw": float(load_str) if load_str else None,
        "total_amount": float(amount_str) if amount_str else None,
    }
    n = (name or "customer").replace(" ", "_")
    m = (month or "report").replace(" ", "_")
    excel_path = f"/tmp/Energybae_Solar_{n}_{m}.xlsx"
    make_excel(bill, {
        "sun_hours": sun_hours, "efficiency": efficiency,
        "tariff_rate": tariff_rate, "cost_per_kwp": cost_per_kwp
    }, calc, excel_path)
    return summary, excel_path


with gr.Blocks(title="Energybae — Solar Load Calculator", theme=gr.themes.Default()) as demo:
    gr.Markdown("# ⚡ Energybae — Solar Load Calculator")
    gr.Markdown("Upload any electricity bill (PDF/image) → AI extracts the data → get solar sizing + Excel report")

    with gr.Row():
        with gr.Column(scale=1):
            gr.Markdown("### 📄 Step 1 — Upload Bill")
            file_input = gr.File(label="Electricity bill (PDF or image)", file_types=[".pdf", ".jpg", ".jpeg", ".png"])

            gr.Markdown("### ⚙️ Solar Parameters")
            sun_hours    = gr.Slider(3.0, 7.0, value=5.5, step=0.1, label="Sun hours / day")
            efficiency   = gr.Slider(60, 95, value=80, step=1,   label="System efficiency (%)")
            tariff_rate  = gr.Slider(3.0, 15.0, value=8.0, step=0.5, label="Electricity rate (₹/unit)")
            cost_per_kwp = gr.Number(value=65000, label="System cost (₹/kWp)")

            extract_btn = gr.Button("🤖 Extract & Calculate", variant="primary")

        with gr.Column(scale=1):
            gr.Markdown("### 📋 Step 2 — Bill Data (editable)")
            status_box   = gr.Textbox(label="Status", interactive=False)
            f_name       = gr.Textbox(label="Consumer name")
            f_consumer   = gr.Textbox(label="Consumer number")
            f_month      = gr.Textbox(label="Billing month")
            f_tariff     = gr.Textbox(label="Tariff category")
            f_units      = gr.Textbox(label="Units consumed (kWh/month)")
            f_load       = gr.Textbox(label="Sanctioned load (kW)")
            f_amount     = gr.Textbox(label="Total bill amount (₹)")

            recalc_btn   = gr.Button("🔄 Recalculate with edited values")

            gr.Markdown("### 📊 Step 3 — Solar Sizing Results")
            results_box  = gr.Textbox(label="Results", lines=6, interactive=False)

            gr.Markdown("### 📥 Step 4 — Download Excel")
            excel_out    = gr.File(label="Download filled Excel report")

    # Wire up
    extract_btn.click(
        fn=process_bill,
        inputs=[file_input, sun_hours, efficiency, tariff_rate, cost_per_kwp],
        outputs=[status_box, f_name, f_consumer, f_month, f_tariff,
                 f_units, f_load, f_amount, results_box, excel_out],
    )

    recalc_btn.click(
        fn=recalculate,
        inputs=[f_units, sun_hours, efficiency, tariff_rate, cost_per_kwp,
                f_name, f_consumer, f_month, f_tariff, f_load, f_amount],
        outputs=[results_box, excel_out],
    )

if __name__ == "__main__":
    demo.launch()